import yfinance as yf
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import os
import datetime
from datetime import date


print('Enter Ticker:')
input = input('').upper()
t = yf.Ticker(input)
balance_sheet = t.balance_sheet
stockInfo = pd.DataFrame(balance_sheet)
stockInfo.to_excel(r'', index=True) #enter in directory to which xlsx file will be stored.

date = date.today()

excel_file_path = r'' #enter in directory to which xlsx file will be stored.

# loads pandas dataframe (df) workbook created with pandas to allow for changes
stock_data = load_workbook(excel_file_path)
writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')
ws1 = stock_data.active
ws1.title = str(input) + ' Yearly Balance Sheet'

dividends = stock_data.create_sheet('Dividends', index=True)
div = pd.DataFrame(t.analysis)
div.to_excel(writer, sheet_name='Dividends')

writer.save()
stock_data.save(excel_file_path)
stock_data.close()


