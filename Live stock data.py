from yahoo_fin.stock_info import *
import yahoo_fin as yf
import numpy as np
import pandas as pd
import openpyxl
import smtplib
import os
from datetime import date
from datetime import datetime
from datetime import timedelta
import time
import datetime
import openpyxl.utils.datetime
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
import pyodbc
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from smtplib import SMTP
import xlsxwriter
import html5lib


def send_email(files, recipients, subject, body):
    emailfrom = ""
    subject = 'Underperforming Large-Cap Stocks'
    body = 'Underperforming Large Cap Stocks'

    msg = MIMEMultipart()
    msg["From"] = emailfrom
    msg["To"] = ",".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, 'plain'))

    filename = excel_file_path
    attachment = open(str(filename), 'rb')

    part = MIMEBase('application', "octet-stream")
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment; filename= " + filename)

    msg.attach(part)
    text = msg.as_string()
    server = SMTP("")
    server.sendmail(emailfrom, recipients, msg.as_string())
    server.quit()


yahoo_data = get_undervalued_large_caps()
date = date.today()
df_upload_to_excel_file = yahoo_data.to_excel('Daily Stock Data ' + str(date) + '.xlsx', index=False)
excel_file_path = ('Daily Stock Data ' + str(date) + '.xlsx')

# loads pandas dataframe (df) workbook created with pandas to allow for changes
stock_data = load_workbook(excel_file_path)
ws1 = stock_data.active
ws1.title = '100 Undervalued Large Caps'

stock_data.save(excel_file_path)
stock_data.close()

# sends email
send_email(excel_file_path, [''], 'Daily Stock Letter', '***')
